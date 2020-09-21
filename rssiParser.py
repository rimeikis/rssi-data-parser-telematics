from datetime import datetime
import statistics
from serial.tools.list_ports import comports
import serial
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.chart import LineChart, Reference, Series

VERSION = "rssiParser_2.0.9"


def setup():
    # FUNCTION FOR SETTING UP SERIAL PORT, IMEI, EXCEL WORKBOOK AND START TIME
    sys.stderr.write('\n--- Available ports:\n')
    ports = []
    for n, (port, desc, hwid) in enumerate(sorted(comports()), 1):
        sys.stderr.write('--- {:2}: {:20} {!r}\n'.format(n, port, desc))
        ports.append(port)

    while True:
        port = input('>>> Enter port index or full name: ')
        try:
            index = int(port) - 1
            if not 0 <= index < len(ports):
                sys.stderr.write('--- Invalid index!\n')
                continue
        except ValueError:
            pass
        else:
            port = ports[index]

        serial_port = serial.Serial()
        serial_port.port = port
        serial_port.baudrate = 4800
        serial_port.bytesize = 5
        serial_port.timeout = 1
        serial_port.stopbits = serial.STOPBITS_ONE

        imei = str(input(">>> Enter target IMEI address: "))
        imei = imei.encode('utf-8')
        process_start_time = datetime.now()
        workbook = Workbook()
        main(serial_port, imei, process_start_time, workbook, 1)


def format_worksheet(worksheet, scenario_note, imei, scenario_start_time, rssi_power_list, rssi_average_list):
    # FUNCTION FOR SETTING UP THE EXCEL WORKSHEET LAYOUT, FORMAT AND CHART
    line_chart = LineChart()
    line_chart.title = "RSSI Power Graph"
    line_chart.y_axis.title = 'RSSI Level'
    line_chart.x_axis.title = 'Data Count'
    rssi_power_data = Reference(worksheet, min_col=9, min_row=8,
                                max_col=9, max_row=len(rssi_power_list)+7)
    power_series = Series(rssi_power_data, title="Actual RSSI Power")
    line_chart.append(power_series)
    rssi_average_data = Reference(worksheet, min_col=15, min_row=8,
                                  max_col=15, max_row=len(rssi_average_list)+7)
    average_series = Series(rssi_average_data, title="Averaged RSSI Power")
    line_chart.append(average_series)
    worksheet.add_chart(line_chart, "Q6")

    worksheet.merge_cells('A1:O1')
    worksheet.merge_cells('A3:O3')
    worksheet.merge_cells('A4:O4')
    worksheet.merge_cells('A6:C6')
    worksheet.merge_cells('E6:I6')
    worksheet.merge_cells('K6:O6')

    worksheet['A3'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['A6'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['A7'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['B7'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['C7'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['E6'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['E7'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['F7'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['G7'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['H7'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['I7'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['K6'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['K7'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['L7'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['M7'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['N7'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    worksheet['O7'].fill = PatternFill(
        start_color="1e90ff", end_color="1e90ff", fill_type="solid")

    worksheet['A1'].alignment = Alignment(horizontal='center')
    worksheet['A3'].alignment = Alignment(horizontal='center')
    worksheet['A4'].alignment = Alignment(horizontal='center')
    worksheet['A6'].alignment = Alignment(horizontal='center')
    worksheet['D6'].alignment = Alignment(horizontal='center')
    worksheet['E6'].alignment = Alignment(horizontal='center')
    worksheet['J6'].alignment = Alignment(horizontal='center')
    worksheet['K6'].alignment = Alignment(horizontal='center')

    worksheet['A3'].font = Font(bold=True)
    worksheet['A6'].font = Font(bold=True)
    worksheet['K6'].font = Font(bold=True)
    worksheet['A1'].font = Font(bold=True)
    worksheet['E6'].font = Font(bold=True)

    worksheet['A1'] = 'RSSI PARSE RESULTS'
    worksheet['A3'] = 'Scenario notes'
    worksheet['A4'] = scenario_note
    worksheet['A6'] = 'Test details'
    worksheet['A7'] = 'Target IMEI'
    worksheet['A8'] = imei
    worksheet['B7'] = 'Start time'
    worksheet['B8'] = scenario_start_time
    worksheet['C7'] = 'Duration'
    worksheet['E6'] = 'Actual RSSI Power'
    worksheet['E7'] = 'Data count'
    worksheet['E8'] = '=COUNT(I8:I1048576)'
    worksheet['F7'] = 'Average'
    worksheet['F8'] = '=AVERAGE(I8:I1048576)'
    worksheet['G7'] = 'Max'
    worksheet['G8'] = '=MAX(I8:I1048576)'
    worksheet['H7'] = 'Min'
    worksheet['H8'] = '=MIN(I8:I1048576)'
    worksheet['I7'] = 'Data'
    worksheet['K6'] = 'Averaged RSSI Power'
    worksheet['K7'] = 'Data count'
    worksheet['K8'] = '=COUNT(O8:O1048576)'
    worksheet['L7'] = 'Average'
    worksheet['L8'] = '=AVERAGE(O8:O1048576)'
    worksheet['M7'] = 'Max'
    worksheet['M8'] = '=MAX(O8:O1048576)'
    worksheet['N7'] = 'Min'
    worksheet['N8'] = '=MIN(O8:O1048576)'
    worksheet['O7'] = 'Data'


def terminal_display(process_start_time, imei, rssi_power, rssi_average, rssi_power_list, rssi_average_list):
    # FUNCTION FOR DISPLAYING LIVE PARSE DATA IN THE TERMINAL
    sys.stderr.write(
        '\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n')
    sys.stderr.write(
        '|-------------------------------------------------|------------------------------------------------|\n')
    sys.stderr.write(
        '|~~~~~~~~~~~~~~~~~~ RSSI DATA ~~~~~~~~~~~~~~~~~~~~|~~~~~~~~~~~~~~~~~~ PARSE DETAILS ~~~~~~~~~~~~~~~|\n')
    sys.stderr.write(
        '|-------------------------------------------------|------------------------------------------------|\n')
    sys.stderr.write('|  {:^20}  |  {:^20}  |  {:^44}  |\n'.format(
        "ACTUAL RSSI POWER", "AVERAGED RSSI POWER", "Target IMEI: " + imei.decode('utf-8')))
    sys.stderr.write(
        '|------------------------|------------------------|------------------------------------------------|\n')
    sys.stderr.write('|  {:^20}  |  {:^20}  |  {:^44}  |\n'.format(
        f"Current: {rssi_power}", f"Current: {rssi_average}", f"Start: {process_start_time}"))
    sys.stderr.write('|  {:^20}  |  {:^20}  |------------------------------------------------|\n'.format(
        f"Count: {len(rssi_power_list)}", f"Count: {len(rssi_average_list)}"))
    sys.stderr.write('|  {:^20}  |  {:^20}  |  {:^40}  |\n'.format(f"Max: {max(rssi_power_list)}",
                                                                   f"Max: {max(rssi_average_list)}", "Hit 'CTRL+C' to stop parsing. A file will be"))
    sys.stderr.write('|  {:^20}  |  {:^20}  |  {:^40}  |\n'.format(f"Min: {min(rssi_power_list)}",
                                                                   f"Min: {min(rssi_average_list)}", "generated containing summary of parsed data."))
    sys.stderr.write('|  {:^20}  |  {:^20}  |  {:^44}  |\n'.format(
        f"Avg: {round(statistics.mean(rssi_power_list))}", f"Avg: {round(statistics.mean(rssi_average_list))}", ""))
    sys.stderr.write(
        '|------------------------|------------------------|------------------------------------------------|\n')
    sys.stderr.write('\n\n\n\n\n\n\n\n\n\n')


def formatted_time():
    # FUNCTION FOR ADJUSTING DATETIME FORMAT TO FIT FILE NAMING RULES
    now = datetime.now()
    date_time_formatted = now.strftime("%Y-%m-%d %H_%M_%S")
    return date_time_formatted


def main(serial_port, imei, process_start_time, workbook, scenario_count):
    # MAIN FUNCTION WHICH PERFORMS RSSI DATA PARSING
    rssi_power_list = []
    rssi_average_list = []
    scenario_number = scenario_count
    scenario_note = input('>>> Enter scenario description: ')
    scenario_start_time = datetime.now()
    worksheet = workbook.create_sheet(title='Scenario ' + str(scenario_number))

    try:
        serial_port.open()
        serial_port.reset_input_buffer()
        while (serial_port.is_open):
            serial_string = serial_port.read_until(b'\r')

            if (imei in serial_string) and (b'Teltonika' in serial_string):
                string_elements_list = serial_string.split()
                rssi_power = int(string_elements_list[9].decode('utf-8')[:-1])
                rssi_average = int(
                    string_elements_list[11].decode('utf-8')[:-1])
                rssi_power_list.append(rssi_power)
                rssi_average_list.append(rssi_average)
                worksheet['I'+str(len(rssi_power_list)+7)] = rssi_power
                worksheet['O'+str(len(rssi_average_list)+7)
                          ] = rssi_average
                terminal_display(process_start_time, imei, rssi_power,
                                 rssi_average, rssi_power_list, rssi_average_list)
            else:
                continue

    except KeyboardInterrupt:
        serial_port.reset_input_buffer()
        serial_port.close()
        worksheet['C8'] = str(datetime.now() - scenario_start_time)
        format_worksheet(worksheet, scenario_note, imei,
                         scenario_start_time, rssi_power_list, rssi_average_list)
        new_scenario = input(
            ">>> Do you wish to continue with another scenario for this test? (Y/N): ")

        if new_scenario.lower() == 'Y'.lower():
            scenario_count = scenario_number + 1
            main(serial_port, imei, process_start_time, workbook, scenario_count)
        else:
            sys.stderr.write('\n--- Parsing terminated!\n')
            sys.stderr.write('--- Process duration: ' +
                             str(datetime.now() - process_start_time) + '\n')
            workbook.remove(workbook['Sheet'])
            workbook.save(filename='parser_results_' +
                          formatted_time() + '.xlsx')
            sys.stderr.write('--- File "parser_results_' + formatted_time() +
                             '.xlsx" containing parsed RSSI data\n--- is next to the ' + VERSION + '.exe" launcher.')

    except IOError:
        sys.stderr.write(
            "IO Error! Possible reason - unable to open port.")
        pass

    except EOFError:
        sys.stderr.write(
            "EOF Error! Failed to read full line from serial port.")
        pass

    finally:
        quit()


setup()
# 359633107038940
# 359633105677517
