## CHANGELOG ##
# sutvarkiau laika, kuris bloga duration ikelia i exceli.
# pridet available MAC adresu sarasa po scan'o. Islistint unikalius IMEI ir ju MAC salia vienas kito. Su 1,2,3... inputais pasirinkt kurio rssi nori detektint

from datetime import datetime
import statistics
from serial.tools.list_ports import comports
import serial
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
VERSION = "rssiParser_2.0.6"


def ask_for_port():
    sys.stderr.write('\n--- Available ports:\n')
    ports = []
    for n, (port, desc, hwid) in enumerate(sorted(comports()), 1):
        sys.stderr.write('--- {:2}: {:20} {!r}\n'.format(n, port, desc))
        ports.append(port)
    while True:
        port = input('>>> Enter port index or full name: ')
        try:
            index = int(port) - 1
            if not 0 <= index < len(ports):
                sys.stderr.write('--- Invalid index!\n')
                continue
        except ValueError:
            pass
        else:
            port = ports[index]
        return port


serialPort = serial.Serial(
    port=ask_for_port(),
    baudrate=115200,
    bytesize=8,
    timeout=2,
    stopbits=serial.STOPBITS_ONE)


def formattedTime():
    now = datetime.now()
    dateTimeFormatted = now.strftime("%Y-%m-%d %H_%M_%S")
    return dateTimeFormatted


def setup():
    mac = str(input(">>> Enter target MAC address: "))
    mac = mac.encode('utf-8')
    wb = Workbook()
    process_start_time = datetime.now()
    function(mac, wb, 1, process_start_time)


def function(mac, wb, x, process_start_time):
    scenarioNumber = x
    note = input('>>> Enter scenario description: ')
    scenario_start_time = datetime.now()
    ws1 = wb.create_sheet(title='Scenario ' + str(scenarioNumber))
    rssiList = []
    ws1.merge_cells('A1:H1')
    ws1.merge_cells('A3:H3')
    ws1.merge_cells('A4:H4')
    ws1['A3'].fill = PatternFill(
        start_color="A59D9E", end_color="A59D9E", fill_type="solid")
    ws1['A6'].fill = PatternFill(
        start_color="A59D9E", end_color="A59D9E", fill_type="solid")
    ws1['B6'].fill = PatternFill(
        start_color="A59D9E", end_color="A59D9E", fill_type="solid")
    ws1['C6'].fill = PatternFill(
        start_color="A59D9E", end_color="A59D9E", fill_type="solid")
    ws1['D6'].fill = PatternFill(
        start_color="A59D9E", end_color="A59D9E", fill_type="solid")
    ws1['E6'].fill = PatternFill(
        start_color="A59D9E", end_color="A59D9E", fill_type="solid")
    ws1['F6'].fill = PatternFill(
        start_color="A59D9E", end_color="A59D9E", fill_type="solid")
    ws1['G6'].fill = PatternFill(
        start_color="A59D9E", end_color="A59D9E", fill_type="solid")
    ws1['H6'].fill = PatternFill(
        start_color="A59D9E", end_color="A59D9E", fill_type="solid")
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1['A1'].font = Font(bold=True)

    ws1['A1'] = 'RSSI PARSE RESULTS'
    ws1['A3'] = 'Scenario notes'
    ws1['A4'] = note
    ws1['A6'] = 'Target MAC'
    ws1['A7'] = mac
    ws1['B6'] = 'Start time'
    ws1['B7'] = scenario_start_time
    ws1['C6'] = 'Duration'
    ws1['D6'] = 'Data count'
    ws1['D7'] = '=COUNT(H7:H1048576)'
    ws1['E6'] = 'Average'
    ws1['E7'] = '=AVERAGE(H7:H1048576)'
    ws1['F6'] = 'Max'
    ws1['F7'] = '=MAX(H7:H1048576)'
    ws1['G6'] = 'Min'
    ws1['G7'] = '=MIN(H7:H1048576)'
    ws1['H6'] = 'Data'

    try:
        while (True):
            if (serialPort.in_waiting > 0):
                serialString = serialPort.read_until(b'\r')

                if mac in serialString:
                    stringElementsList = serialString.split()
                    for value in stringElementsList:
                        if value.startswith(b'-') and value != (b'->'):
                            rssiList.append(int(value))
                            ws1['H'+str(len(rssiList)+6)] = int(value)

                            print("===============================")
                            print(serialString)
                            sys.stderr.write(
                                '\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n')
                            sys.stderr.write(
                                '|------------------------------------------------|------------------------------------------------|\n')
                            sys.stderr.write(
                                '|~~~~~~~~~~~~~~~~~~ RSSI DATA ~~~~~~~~~~~~~~~~~~~|~~~~~~~~~~~~~~~~~~ PARSE DETAILS ~~~~~~~~~~~~~~~|\n')
                            sys.stderr.write(
                                '|------------------------------------------------|------------------------------------------------|\n')
                            sys.stderr.write('|    {:^40}    |    {:^40}    |\n'.format(
                                "", "Start: " + str(process_start_time)))
                            sys.stderr.write('|    {:^40}    |------------------------------------------------|\n'.format(
                                "Current: " + value.decode('utf-8') + 'dBm'))
                            sys.stderr.write('|    {:^40}    |    {:^40}    |\n'.format(
                                "", "Target MAC: " + mac.decode('utf-8')))
                            sys.stderr.write(
                                '|------------------------------------------------|------------------------------------------------|\n')
                            sys.stderr.write('|  {:^20}  | {:^20}  |  {:^40}  |\n'.format("Count: " + str(len(rssiList)), "Avg: " + str(
                                round(statistics.mean(rssiList))), "Hit 'CTRL+C' to stop parsing. A file will be"))
                            sys.stderr.write(
                                '|------------------------------------------------|{:>49}\n'.format('|'))
                            sys.stderr.write('|  {:^20}  | {:^20}  |  {:^40}  |\n'.format(
                                "Max: " + str(max(rssiList)), "Min: " + str(min(rssiList)), "generated containing summary of parsed data."))
                            sys.stderr.write(
                                '|------------------------------------------------|------------------------------------------------|')
                            sys.stderr.write('\n\n\n\n\n\n\n\n\n\n')
    except KeyboardInterrupt:
        ws1['C7'] = str(datetime.now() - scenario_start_time)
        ask = input(
            ">>> Do you wish to continue with another scenario for this test? (Y/N): ")

        if ask.lower() == 'Y'.lower():
            x = scenarioNumber + 1
            function(mac, wb, x, process_start_time)
        else:
            sys.stderr.write('\n--- Parsing terminated!\n')
            sys.stderr.write('--- Process duration: ' +
                             str(datetime.now() - process_start_time) + '\n')

            wb.remove(wb['Sheet'])

            wb.save(filename='parser_results_' + formattedTime() + '.xlsx')
            sys.stderr.write('--- File "parser_results_' + formattedTime() +
                             '.xlsx" containing parsed RSSI data\n--- is next to the ' + VERSION + '.exe" launcher.')
    finally:
        input()


setup()
